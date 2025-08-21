import os
import uuid
import io
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

from PIL import Image
import easyocr
import openpyxl

from models import OcrResponse, GenerateCardRequest, GenerateCardResponse
from parser import parse_recipe

# --- Kaustad ---
BASE_DIR = os.path.dirname(__file__)
STORAGE_DIR = os.path.join(BASE_DIR, "storage")
IMAGES_DIR = os.path.join(STORAGE_DIR, "images")
TEXTS_DIR = os.path.join(STORAGE_DIR, "texts")
RECIPES_XLSX = os.path.join(STORAGE_DIR, "recipes.xlsx")
TEMPLATES_DIR = os.path.join(BASE_DIR, "excel_templates")
TEMPLATE_XLSX = os.path.join(TEMPLATES_DIR, "tech_card_template.xlsx")

os.makedirs(IMAGES_DIR, exist_ok=True)
os.makedirs(TEXTS_DIR, exist_ok=True)

app = FastAPI(title="OCR Recipe API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Lae easyocr Reader ühe korra (eesti+inglise). GPU=False vaikimisi CPU jaoks.
reader = easyocr.Reader(["et", "en"], gpu=False)

# --- Abifunktsioonid ---

def init_recipes_xlsx():
    if not os.path.exists(RECIPES_XLSX):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "recipes"
        ws.append(["id", "timestamp", "title", "text", "image_path", "text_path"])
        wb.save(RECIPES_XLSX)

init_recipes_xlsx()


def append_recipe_row(rid: str, title: str, text: str, image_path: str, text_path: str):
    wb = openpyxl.load_workbook(RECIPES_XLSX)
    ws = wb["recipes"]
    ws.append([rid, datetime.utcnow().isoformat(), title, text, image_path, text_path])
    wb.save(RECIPES_XLSX)


def extract_text_from_image_bytes(img_bytes: bytes) -> str:
    # PIL avamine + vajadusel rota/kontrasti eel-töötlus
    image = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    # OCR
    results = reader.readtext(np.array(image), detail=0)  # ainult tekstijupid
    text = "\n".join(results)
    return text

# easyocr vajab numpy
import numpy as np

# --- Endpointid ---

@app.post("/ocr", response_model=OcrResponse)
async def ocr_image(file: UploadFile = File(...)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Palun laadi üles pildifail.")

    rid = str(uuid.uuid4())[:8]
    image_ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
    image_path = os.path.join(IMAGES_DIR, f"{rid}{image_ext}")

    # salvesta originaalpilt
    raw = await file.read()
    with open(image_path, "wb") as f:
        f.write(raw)

    # OCR
    text = extract_text_from_image_bytes(raw)
    parsed = parse_recipe(text)
    title = parsed.get("title") or "Retsept"

    # salvesta tekst
    text_path = os.path.join(TEXTS_DIR, f"{rid}.txt")
    with open(text_path, "w", encoding="utf-8") as f:
        f.write(text)

    # lisa master Excelisse
    append_recipe_row(rid, title, text, image_path, text_path)

    return OcrResponse(id=rid, text=text, title=title, saved_text_path=text_path)


@app.post("/generate-tech-card", response_model=GenerateCardResponse)
async def generate_tech_card(req: GenerateCardRequest):
    # tekst tuleb kas id alusel master-failist või req.text väljast
    rid = req.id or str(uuid.uuid4())[:8]

    if req.id and not req.text:
        # loe tekst recipes.xlsx-ist
        if not os.path.exists(RECIPES_XLSX):
            raise HTTPException(status_code=404, detail="recipes.xlsx puudub")
        wb = openpyxl.load_workbook(RECIPES_XLSX)
        ws = wb["recipes"]
        text = None
        title = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == req.id:
                title = row[2]
                text = row[3]
                break
        if text is None:
            raise HTTPException(status_code=404, detail=f"ID {req.id} ei leitud")
    else:
        text = req.text or ""
        title = parse_recipe(text).get("title") or "Retsept"

    parsed = parse_recipe(text)

    # Täida template
    if not os.path.exists(TEMPLATE_XLSX):
        raise HTTPException(status_code=500, detail="Template puudub")

    wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    ws = wb.active  # eeldame, et vorm on esimesel lehel

    # --- MAPPIMINE ---
    # Kohanda need lahtrid oma template'ile vastavaks!
    mapping = {
        "B2": parsed.get("title", ""),
        "B5": parsed.get("ingredients", ""),
        "B12": parsed.get("steps", ""),
    }

    for cell, value in mapping.items():
        ws[cell] = value

    # Salvesta uue failina
    out_path = os.path.join(STORAGE_DIR, f"tech_card_{rid}.xlsx")
    wb.save(out_path)

    return GenerateCardResponse(id=rid, tech_card_path=out_path, title=parsed.get("title", ""))


@app.get("/download-tech-card/{rid}")
async def download_tech_card(rid: str):
    path = os.path.join(STORAGE_DIR, f"tech_card_{rid}.xlsx")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Faili ei leitud")
    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=os.path.basename(path))
